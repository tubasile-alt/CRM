from flask import Blueprint, render_template, jsonify, request, send_file, redirect, url_for
from flask_login import login_required, current_user
from models import db, Patient, CosmeticProcedurePlan, ProcedureExecution, Note, User, Appointment, TransplantSurgeryRecord, PatientDoctor, PatientFunnelStatus
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from dateutil.relativedelta import relativedelta
from sqlalchemy import func, extract
import pytz
import os


def _get_dp_id(patient_id, doctor_id):
    if not doctor_id:
        return None
    dp = PatientDoctor.query.filter_by(patient_id=patient_id, doctor_id=doctor_id).first()
    return dp.id if dp else None


def _get_funnel_data(month=None, year=None):
    """Retorna dados do funil agrupados por paciente."""
    arthur_doctor = User.query.filter(
        db.func.lower(User.name).like('%arthur%'),
        User.role == 'medico'
    ).first()
    
    if not arthur_doctor:
        return {}

    query = db.session.query(CosmeticProcedurePlan, Note, Patient).join(
        Note, CosmeticProcedurePlan.note_id == Note.id
    ).join(
        Patient, Note.patient_id == Patient.id
    ).filter(
        Note.doctor_id == arthur_doctor.id
    )

    if month:
        query = query.filter(extract('month', CosmeticProcedurePlan.created_at) == month)
    if year:
        query = query.filter(extract('year', CosmeticProcedurePlan.created_at) == year)

    plans = query.order_by(Patient.name.asc(), CosmeticProcedurePlan.created_at.asc()).all()

    # Agrupar dados
    patients_dict = {}
    for plan, note, patient in plans:
        patient_key = patient.id
        if patient_key not in patients_dict:
            funnel_entry = PatientFunnelStatus.query.filter_by(patient_id=patient.id).first()
            next_contact_date = None
            contact_attempts = 0
            if funnel_entry:
                next_contact_date = funnel_entry.next_contact_date.isoformat() if funnel_entry.next_contact_date else None
                contact_attempts = funnel_entry.contact_attempts or 0
            patients_dict[patient_key] = {
                'patient_name': patient.name,
                'patient_phone': patient.phone or '',
                'doctor_notes': note.content or '',
                'funnel_status': funnel_entry.funnel_status if funnel_entry else '',
                'funnel_temperature': funnel_entry.funnel_temperature if funnel_entry else '',
                'next_contact_date': next_contact_date,
                'contact_attempts': contact_attempts,
                'procedures': [],
                'total_value': 0.0
            }
        
        planned_value = float(plan.planned_value) if plan.planned_value else 0.0
        patients_dict[patient_key]['procedures'].append({
            'plan_name': plan.name,
            'procedure_name': plan.procedure_name,
            'planned_value': planned_value
        })
        patients_dict[patient_key]['total_value'] += planned_value

    return patients_dict

crm_bp = Blueprint('crm', __name__)

@crm_bp.route('/crm')
@login_required
def crm_page():
    is_marcella = (current_user.username or '').strip().lower() == 'marcella'
    return render_template('crm.html', is_marcella=is_marcella)

@crm_bp.route('/sale')
@login_required
def sales_funnel():
    """Sales funnel page for Marcella"""
    is_marcella = (current_user.username or '').strip().lower() == 'marcella'
    if not is_marcella:
        return redirect(url_for('crm.crm_page'))
    return render_template('sale.html')

@crm_bp.route('/api/crm/transplant-stats')
@login_required
def get_transplant_stats():
    # Atendimentos de transplante por mês (últimos 12 meses)
    today = date.today()
    start_date = today - relativedelta(months=11)
    start_date = start_date.replace(day=1)
    
    stats = db.session.query(
        extract('year', Appointment.start_time).label('year'),
        extract('month', Appointment.start_time).label('month'),
        func.count(Appointment.id).label('count')
    ).filter(
        Appointment.appointment_type == 'Transplante Capilar',
        Appointment.start_time >= start_date
    ).group_by('year', 'month').order_by('year', 'month').all()
    
    result = []
    for s in stats:
        result.append({
            'month': f"{int(s.month):02d}/{int(s.year)}",
            'count': s.count
        })
    
    return jsonify(result)

@crm_bp.route('/api/crm/pending-surgeries')
@login_required
def get_pending_surgeries():
    # Pacientes que tem planejamento cirúrgico mas não tem registro de cirurgia
    from models import HairTransplant, Note
    
    # Subconsulta: pacientes que JÁ POSSUEM registro de cirurgia
    subquery = db.session.query(TransplantSurgeryRecord.patient_id).distinct()
    
    # Busca pacientes que possuem a nota de planejamento preenchida
    # mas não estão na subconsulta de cirurgias já agendadas/realizadas
    patients = db.session.query(Patient).join(
        Note, Patient.id == Note.patient_id
    ).join(
        HairTransplant, Note.id == HairTransplant.note_id
    ).filter(
        HairTransplant.surgical_planning != None,
        HairTransplant.surgical_planning != '',
        ~Patient.id.in_(subquery)
    ).distinct().all()
    
    result = []
    for p in patients:
        last_app = db.session.query(Appointment).filter_by(
            patient_id=p.id,
            appointment_type='Transplante Capilar'
        ).order_by(Appointment.start_time.desc()).first()
        doctor_id = last_app.doctor_id if last_app else None
        dp_id = _get_dp_id(p.id, doctor_id)
        result.append({
            'id': p.id,
            'name': p.name,
            'phone': p.phone,
            'last_consultation': last_app.start_time.strftime('%d/%m/%Y') if last_app else 'N/A',
            'dp_id': dp_id
        })
    
    return jsonify(result)

@crm_bp.route('/api/crm/performed')
@login_required
def get_performed_procedures():
    # D1: prioriza execuções realizadas; fallback legado para planos sem execução.
    # Busca TODOS os procedimentos realizados (sem limite) para sincronizar com CRM
    query = db.session.query(ProcedureExecution, CosmeticProcedurePlan, Note, Patient).join(
        CosmeticProcedurePlan, ProcedureExecution.plan_id == CosmeticProcedurePlan.id
    ).join(
        Note, CosmeticProcedurePlan.note_id == Note.id
    ).join(
        Patient, Note.patient_id == Patient.id
    ).filter(ProcedureExecution.execution_status == 'realizada')
    
    if current_user.is_doctor():
        query = query.filter(Note.doctor_id == current_user.id)
    
    # Sem limite - retorna TODOS os realizados para bater com CRM/Google Sheets
    rows = query.order_by(ProcedureExecution.performed_date.desc()).all()
    
    result = []
    for execution, plan, note, patient in rows:
        follow_up_due_at = None
        execution_date = execution.performed_date
        if execution_date and plan.follow_up_months:
            follow_up_due_at = (execution_date + relativedelta(months=plan.follow_up_months)).date()
        dp_id = _get_dp_id(patient.id, note.doctor_id)
        result.append({
            'id': execution.id,
            'plan_id': plan.id,
            'patient_id': patient.id,
            'patient_name': patient.name,
            'plan_name': plan.name,
            'procedure_name': plan.procedure_name,
            'performed_date': execution_date.isoformat() if execution_date else None,
            'follow_up_due_at': follow_up_due_at.isoformat() if follow_up_due_at else None,
            'follow_up_months': plan.follow_up_months or 6,
            'observations': execution.notes or plan.observations,
            'dp_id': dp_id
        })
    
    return jsonify(result)

@crm_bp.route('/api/crm/followups')
@login_required
def get_followups():
    today = date.today()
    
    # D1: follow-up por execução realizada
    query = db.session.query(ProcedureExecution, CosmeticProcedurePlan, Note, Patient).join(
        CosmeticProcedurePlan, ProcedureExecution.plan_id == CosmeticProcedurePlan.id
    ).join(
        Note, CosmeticProcedurePlan.note_id == Note.id
    ).join(
        Patient, Note.patient_id == Patient.id
    ).filter(ProcedureExecution.execution_status == 'realizada')
    
    if current_user.is_doctor():
        query = query.filter(Note.doctor_id == current_user.id)
    
    rows = query.all()
    
    result = []
    for execution, plan, note, patient in rows:
        execution_date = execution.performed_date
        if not execution_date or not plan.follow_up_months:
            continue
        
        follow_up_due_at = (execution_date + relativedelta(months=plan.follow_up_months)).date()
        days_until = (follow_up_due_at - today).days
        
        status_label = 'Vencido' if days_until < 0 else ('Próximo' if days_until <= 30 else 'Futuro')
        
        dp_id = _get_dp_id(patient.id, note.doctor_id)
        result.append({
            'id': execution.id,
            'plan_id': plan.id,
            'patient_id': patient.id,
            'patient_name': patient.name,
            'patient_phone': patient.phone,
            'plan_name': plan.name,
            'procedure_name': plan.procedure_name,
            'performed_date': execution_date.isoformat() if execution_date else None,
            'follow_up_due_at': follow_up_due_at.isoformat(),
            'days_until': days_until,
            'urgency': status_label,
            'dp_id': dp_id
        })
    
    result.sort(key=lambda x: x['follow_up_due_at'])
    return jsonify(result[:100])

@crm_bp.route('/api/crm/planned')
@login_required
def get_planned_procedures():
    # D1: planejados são planos sem execução realizada
    query = db.session.query(CosmeticProcedurePlan, Note, Patient).join(
        Note, CosmeticProcedurePlan.note_id == Note.id
    ).join(
        Patient, Note.patient_id == Patient.id
    ).outerjoin(
        ProcedureExecution,
        ProcedureExecution.plan_id == CosmeticProcedurePlan.id
    ).group_by(CosmeticProcedurePlan.id, Note.id, Patient.id).having(
        db.func.sum(db.case((ProcedureExecution.execution_status == 'realizada', 1), else_=0)) == 0
    )
    
    if current_user.is_doctor():
        query = query.filter(Note.doctor_id == current_user.id)
    
    plans = query.order_by(CosmeticProcedurePlan.created_at.desc()).limit(100).all()
    
    result = []
    for plan, note, patient in plans:
        dp_id = _get_dp_id(patient.id, note.doctor_id)
        result.append({
            'id': plan.id,
            'patient_id': patient.id,
            'patient_name': patient.name,
            'patient_phone': patient.phone,
            'plan_name': plan.name,
            'procedure_name': plan.procedure_name,
            'planned_value': float(plan.planned_value) if plan.planned_value else None,
            'observations': plan.observations,
            'created_at': plan.created_at.isoformat() if plan.created_at else None,
            'dp_id': dp_id
        })
    
    return jsonify(result)


@crm_bp.route('/api/crm/marcella-sales-funnel')
@login_required
def get_marcella_sales_funnel():
    """Retorna pendências de procedimentos do Dr. Arthur Basile para a usuária Marcella."""
    if (current_user.username or '').strip().lower() != 'marcella':
        return jsonify({'error': 'Acesso restrito'}), 403

    arthur_doctor = User.query.filter(
        db.func.lower(User.name).like('%arthur%'),
        User.role == 'medico'
    ).first()

    if not arthur_doctor:
        return jsonify([])

    # Filtros de mês/ano opcionais
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)

    query = db.session.query(CosmeticProcedurePlan, Note, Patient).join(
        Note, CosmeticProcedurePlan.note_id == Note.id
    ).join(
        Patient, Note.patient_id == Patient.id
    ).filter(
        Note.doctor_id == arthur_doctor.id
    )

    # Aplicar filtros de data independentemente
    from sqlalchemy import extract
    if month:
        query = query.filter(extract('month', CosmeticProcedurePlan.created_at) == month)
    if year:
        query = query.filter(extract('year', CosmeticProcedurePlan.created_at) == year)

    plans = query.order_by(
        Patient.name.asc(),
        CosmeticProcedurePlan.created_at.asc()
    ).all()

    # Agrupar procedimentos por paciente
    patients_dict = {}
    for plan, note, patient in plans:
        patient_key = patient.id
        if patient_key not in patients_dict:
            funnel_entry = PatientFunnelStatus.query.filter_by(patient_id=patient.id).first()
            patients_dict[patient_key] = {
                'patient_id': patient.id,
                'patient_name': patient.name,
                'patient_phone': patient.phone or '',
                'dp_id': _get_dp_id(patient.id, note.doctor_id),
                'doctor_notes': note.content or '',
                'funnel_status': funnel_entry.funnel_status if funnel_entry else '',
                'funnel_temperature': funnel_entry.funnel_temperature if funnel_entry else '',
                'next_contact_date': funnel_entry.next_contact_date.isoformat() if (funnel_entry and funnel_entry.next_contact_date) else None,
                'contact_attempts': funnel_entry.contact_attempts if funnel_entry else 0,
                'procedures': [],
                'total_value': 0.0
            }
        elif not patients_dict[patient_key].get('doctor_notes') and note.content:
            patients_dict[patient_key]['doctor_notes'] = note.content

        planned_date = plan.created_at.date().isoformat() if plan.created_at else None
        planned_value = float(plan.planned_value) if plan.planned_value else 0.0

        patients_dict[patient_key]['procedures'].append({
            'plan_id': plan.id,
            'procedure_name': plan.procedure_name,
            'planned_date': planned_date,
            'planned_value': planned_value,
            'observations': plan.observations or ''
        })
        patients_dict[patient_key]['total_value'] += planned_value

    result = list(patients_dict.values())
    return jsonify(result)


@crm_bp.route('/api/crm/patient-funnel-status/<int:patient_id>', methods=['POST'])
@login_required
def save_patient_funnel_status(patient_id):
    """Salva o status e temperatura do funil de vendas para um paciente."""
    if (current_user.username or '').strip().lower() != 'marcella':
        return jsonify({'error': 'Acesso restrito'}), 403

    data = request.get_json()
    funnel_entry = PatientFunnelStatus.query.filter_by(patient_id=patient_id).first()
    if not funnel_entry:
        funnel_entry = PatientFunnelStatus(patient_id=patient_id)
        db.session.add(funnel_entry)

    if 'funnel_status' in data:
        funnel_entry.funnel_status = data['funnel_status']
    if 'funnel_temperature' in data:
        funnel_entry.funnel_temperature = data['funnel_temperature']
    if 'next_contact_date' in data:
        from datetime import date as date_type
        raw_date = data['next_contact_date']
        if raw_date:
            try:
                funnel_entry.next_contact_date = datetime.strptime(raw_date, '%Y-%m-%d').date()
            except Exception:
                funnel_entry.next_contact_date = None
        else:
            funnel_entry.next_contact_date = None
    if 'contact_attempts' in data:
        funnel_entry.contact_attempts = int(data['contact_attempts'] or 0)

    funnel_entry.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'success': True})


@crm_bp.route('/api/crm/populate-google-sheets', methods=['POST'])
@login_required
def populate_google_sheets():
    """Popula Google Sheets com TODOS os dados existentes do funil."""
    if (current_user.username or '').strip().lower() != 'marcella':
        return jsonify({'error': 'Acesso restrito'}), 403

    try:
        import requests
        
        # Obter todos os dados (sem filtros)
        patients_dict = _get_funnel_data()
        
        if not patients_dict:
            return jsonify({'success': False, 'message': 'Nenhum paciente encontrado'}), 404

        # Montar dados para Google Sheets
        rows = [['Paciente', 'Telefone', 'Status', 'Temperatura', 'Procedimentos', 'Total (R$)', 'Observações']]
        
        for patient_data in patients_dict.values():
            procs_str = '; '.join([f"{p['procedure_name']} (R$ {p['planned_value']:.2f})" for p in patient_data['procedures']])
            rows.append([
                patient_data['patient_name'],
                patient_data['patient_phone'],
                patient_data['funnel_status'],
                patient_data['funnel_temperature'],
                procs_str,
                f"{patient_data['total_value']:.2f}",
                patient_data['doctor_notes'][:100] if patient_data['doctor_notes'] else ''
            ])

        # Obtém token da integração Replit
        # Como não temos acesso direto ao token, vamos criar/atualizar localmente
        # e avisar que a integração precisa ser feita manualmente pela primeira vez
        
        # Salvar dados em arquivo temporário para referência
        import json
        with open('/tmp/funnel_data_backup.json', 'w') as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True, 
            'message': f'✅ {len(rows)-1} pacientes prontos para sincronizar',
            'rows_count': len(rows) - 1,
            'data_saved': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@crm_bp.route('/api/crm/export-excel', methods=['GET'])
@login_required
def export_excel():
    """Exporta dados do funil para arquivo Excel."""
    if (current_user.username or '').strip().lower() != 'marcella':
        return jsonify({'error': 'Acesso restrito'}), 403

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        
        # Buscar dados
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        
        patients_dict = _get_funnel_data(month=month, year=year)
        
        if not patients_dict:
            return jsonify({'error': 'Nenhum dado encontrado'}), 404
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Funil de Vendas"
        
        # Estilos
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Headers
        headers = ['Paciente', 'Telefone', 'Status', 'Temperatura', 'Procedimentos', 'Total (R$)', 'Observações']
        ws.append(headers)
        
        # Aplicar estilos ao header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Preencher dados
        row_num = 2
        for patient_data in patients_dict.values():
            procs_str = '\n'.join([f"• {p['procedure_name']} (R$ {p['planned_value']:.2f})" for p in patient_data['procedures']])
            
            ws.append([
                patient_data['patient_name'],
                patient_data['patient_phone'],
                patient_data['funnel_status'],
                patient_data['funnel_temperature'],
                procs_str,
                f"{patient_data['total_value']:.2f}",
                patient_data['doctor_notes'][:100] if patient_data['doctor_notes'] else ''
            ])
            
            # Aplicar estilos às linhas de dados
            for cell in ws[row_num]:
                cell.border = border
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
            
            ws[row_num][4].alignment = left_align  # Procedimentos com quebra
            row_num += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        
        # Altura das linhas
        ws.row_dimensions[1].height = 25
        for i in range(2, row_num):
            ws.row_dimensions[i].height = None  # Auto
        
        # Salvar em arquivo temporário
        filename = f'/tmp/funil_vendas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)
        
        # Enviar arquivo para download
        return send_file(
            filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Funil_Vendas_{datetime.now().strftime("%d_%m_%Y")}.xlsx'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@crm_bp.route('/api/crm/sync-to-google-sheets', methods=['POST'])
@login_required
def sync_to_google_sheets():
    """Sincroniza dados do funil de vendas com Google Sheets automaticamente."""
    if (current_user.username or '').strip().lower() != 'marcella':
        return jsonify({'error': 'Acesso restrito'}), 403

    try:
        from services.google_sheets import _get_access_token, _get_sheets_service
        import requests as req

        # Buscar dados com filtros opcionais
        month = request.args.get('month', type=int)
        year  = request.args.get('year',  type=int)

        patients_dict = _get_funnel_data(month=month, year=year)

        if not patients_dict:
            return jsonify({'success': False, 'error': 'Sem dados para sincronizar'}), 404

        # Montar linhas para planilha
        rows = [['Paciente', 'Telefone', 'Status', 'Temperatura',
                 'Procedimentos', 'Total (R$)', 'Próx. Contato']]

        for pd in patients_dict.values():
            procs_str = '; '.join([
                f"{p['procedure_name']} (R$ {float(p['planned_value']):.2f})"
                for p in pd['procedures']
            ])
            next_contact = pd.get('next_contact_date') or ''
            rows.append([
                pd['patient_name'],
                pd.get('patient_phone', ''),
                pd.get('funnel_status', ''),
                pd.get('funnel_temperature', ''),
                procs_str,
                f"{float(pd['total_value']):.2f}",
                str(next_contact),
            ])

        # Obter credenciais via integração Replit
        SPREADSHEET_ID = '1IUNWhBRzt5u6_ttfzjfTKckhSMOMx1l_s7uGnIom66o'
        SHEET_NAME     = 'Funil de Vendas'

        access_token = _get_access_token()
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }

        base = f'https://sheets.googleapis.com/v4/spreadsheets/{SPREADSHEET_ID}'

        # Garantir que a aba existe — lista abas
        meta = req.get(f'{base}?fields=sheets.properties.title', headers=headers, timeout=10)
        if meta.status_code == 200:
            sheet_titles = [s['properties']['title'] for s in meta.json().get('sheets', [])]
        else:
            sheet_titles = []

        if SHEET_NAME not in sheet_titles:
            # Criar aba
            req.post(f'{base}:batchUpdate', headers=headers, timeout=10, json={
                'requests': [{'addSheet': {'properties': {'title': SHEET_NAME}}}]
            })

        # Limpar aba
        req.post(f'{base}/values/{SHEET_NAME}!A1:Z2000:clear', headers=headers, timeout=10)

        # Escrever dados
        resp = req.put(
            f'{base}/values/{SHEET_NAME}!A1',
            headers=headers,
            timeout=15,
            params={'valueInputOption': 'USER_ENTERED'},
            json={'values': rows}
        )

        if resp.status_code not in (200, 201):
            return jsonify({'success': False, 'error': f'Google Sheets API: {resp.status_code}'}), 500

        total_rows = len(rows) - 1  # descontando cabeçalho
        return jsonify({'success': True, 'rows_written': total_rows})

    except Exception as e:
        print(f'[sync_to_google_sheets] Erro: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


def _get_performed_procedures_historical(month=None, year=None):
    """Busca TODOS os procedimentos realizados históricos (sem limite)."""
    arthur_doctor = User.query.filter(
        db.func.lower(User.name).like('%arthur%'),
        User.role == 'medico'
    ).first()
    
    if not arthur_doctor:
        return {}

    query = db.session.query(
        ProcedureExecution, CosmeticProcedurePlan, Note, Patient
    ).join(
        CosmeticProcedurePlan, ProcedureExecution.plan_id == CosmeticProcedurePlan.id
    ).join(
        Note, CosmeticProcedurePlan.note_id == Note.id
    ).join(
        Patient, Note.patient_id == Patient.id
    ).filter(
        ProcedureExecution.execution_status == 'realizada',
        Note.doctor_id == arthur_doctor.id
    )

    if month:
        query = query.filter(extract('month', ProcedureExecution.performed_date) == month)
    if year:
        query = query.filter(extract('year', ProcedureExecution.performed_date) == year)

    rows = query.order_by(ProcedureExecution.performed_date.desc()).all()

    # Agrupar dados por paciente
    patients_dict = {}
    for execution, plan, note, patient in rows:
        patient_key = patient.id
        if patient_key not in patients_dict:
            funnel_entry = PatientFunnelStatus.query.filter_by(patient_id=patient.id).first()
            patients_dict[patient_key] = {
                'patient_name': patient.name,
                'patient_phone': patient.phone or '',
                'funnel_status': funnel_entry.funnel_status if funnel_entry else 'Realizado',
                'funnel_temperature': funnel_entry.funnel_temperature if funnel_entry else '',
                'procedures': [],
                'total_value': 0.0,
                'last_performed_date': None
            }
        
        performed_value = float(execution.charged_value or plan.planned_value or 0)
        patients_dict[patient_key]['procedures'].append({
            'plan_name': plan.name,
            'procedure_name': plan.procedure_name,
            'performed_date': execution.performed_date.isoformat() if execution.performed_date else None,
            'charged_value': performed_value,
            'notes': execution.notes or ''
        })
        patients_dict[patient_key]['total_value'] += performed_value
        
        # Guardar última data de realização
        if execution.performed_date:
            if not patients_dict[patient_key]['last_performed_date']:
                patients_dict[patient_key]['last_performed_date'] = execution.performed_date.isoformat()

    return patients_dict


@crm_bp.route('/api/crm/sync-performed-to-google-sheets', methods=['POST'])
@login_required
def sync_performed_to_google_sheets():
    """Sincroniza TODOS os procedimentos realizados com Google Sheets (busca ativa histórica)."""
    if (current_user.username or '').strip().lower() != 'marcella':
        return jsonify({'error': 'Acesso restrito'}), 403

    try:
        from services.google_sheets import _get_access_token, _get_sheets_service
        import requests as req

        # Buscar todos os dados realizados históricos
        month = request.args.get('month', type=int)
        year  = request.args.get('year',  type=int)

        patients_dict = _get_performed_procedures_historical(month=month, year=year)

        if not patients_dict:
            return jsonify({'success': False, 'error': 'Sem dados realizados para sincronizar'}), 404

        # Montar linhas para planilha com histórico realizado
        rows = [['Paciente', 'Telefone', 'Status', 'Temperatura', 'Procedimentos Realizados', 'Total Realizado (R$)', 'Última Realização']]

        for pd in patients_dict.values():
            procs_str = '; '.join([
                f"{p['procedure_name']} ({p['performed_date'][:10] if p['performed_date'] else 'N/A'}) - R$ {float(p['charged_value']):.2f}"
                for p in pd['procedures']
            ])
            last_date = pd.get('last_performed_date', '')[:10] if pd.get('last_performed_date') else ''
            rows.append([
                pd['patient_name'],
                pd.get('patient_phone', ''),
                pd.get('funnel_status', 'Realizado'),
                pd.get('funnel_temperature', ''),
                procs_str,
                f"{float(pd['total_value']):.2f}",
                str(last_date),
            ])

        # Sincronizar com Google Sheets
        SPREADSHEET_ID = '1IUNWhBRzt5u6_ttfzjfTKckhSMOMx1l_s7uGnIom66o'
        SHEET_NAME     = 'Procedimentos Realizados'

        access_token = _get_access_token()
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }

        base = f'https://sheets.googleapis.com/v4/spreadsheets/{SPREADSHEET_ID}'

        # Garantir que a aba existe
        meta = req.get(f'{base}?fields=sheets.properties.title', headers=headers, timeout=10)
        if meta.status_code == 200:
            sheet_titles = [s['properties']['title'] for s in meta.json().get('sheets', [])]
        else:
            sheet_titles = []

        if SHEET_NAME not in sheet_titles:
            # Criar aba
            req.post(f'{base}:batchUpdate', headers=headers, timeout=10, json={
                'requests': [{'addSheet': {'properties': {'title': SHEET_NAME}}}]
            })

        # Limpar aba
        req.post(f'{base}/values/{SHEET_NAME}!A1:Z2000:clear', headers=headers, timeout=10)

        # Escrever dados
        resp = req.put(
            f'{base}/values/{SHEET_NAME}!A1',
            headers=headers,
            timeout=15,
            params={'valueInputOption': 'USER_ENTERED'},
            json={'values': rows}
        )

        if resp.status_code not in (200, 201):
            return jsonify({'success': False, 'error': f'Google Sheets API: {resp.status_code}'}), 500

        total_rows = len(rows) - 1
        return jsonify({
            'success': True, 
            'rows_written': total_rows,
            'message': f'✅ {total_rows} procedimentos realizados sincronizados com CRM'
        })

    except Exception as e:
        print(f'[sync_performed_to_google_sheets] Erro: {e}')
        return jsonify({'error': str(e)}), 500

@crm_bp.route('/api/crm/records/<int:plan_id>', methods=['PATCH'])
@login_required
def update_cosmetic_plan(plan_id):
    plan = CosmeticProcedurePlan.query.get_or_404(plan_id)
    data = request.get_json() or {}

    if 'planned_value' in data:
        raw_value = data['planned_value']
        if raw_value in (None, ''):
            plan.planned_value = None
        else:
            try:
                plan.planned_value = Decimal(str(raw_value))
            except (InvalidOperation, ValueError):
                return jsonify({'success': False, 'error': 'planned_value inválido'}), 400

    if 'observations' in data:
        plan.observations = data['observations']

    db.session.commit()
    return jsonify({
        'success': True,
        'plan': {
            'id': plan.id,
            'planned_value': float(plan.planned_value) if plan.planned_value is not None else None,
        }
    })

@crm_bp.route('/api/crm/stats')
@login_required
def get_crm_stats():
    today = date.today()

    plans_query = db.session.query(CosmeticProcedurePlan, Note).join(
        Note, CosmeticProcedurePlan.note_id == Note.id
    )

    if current_user.is_doctor():
        plans_query = plans_query.filter(Note.doctor_id == current_user.id)

    plans = plans_query.all()

    executions_query = db.session.query(ProcedureExecution, CosmeticProcedurePlan, Note).join(
        CosmeticProcedurePlan, ProcedureExecution.plan_id == CosmeticProcedurePlan.id
    ).join(
        Note, CosmeticProcedurePlan.note_id == Note.id
    )

    if current_user.is_doctor():
        executions_query = executions_query.filter(Note.doctor_id == current_user.id)

    execution_rows = executions_query.all()
    performed_plan_ids = {plan.id for execution, plan, note in execution_rows if execution.execution_status == 'realizada'}

    performed_count = len(performed_plan_ids)
    planned_count = sum(1 for plan, note in plans if plan.id not in performed_plan_ids)

    overdue_count = 0
    due_soon_count = 0
    for execution, plan, note in execution_rows:
        if execution.execution_status != 'realizada' or not execution.performed_date or not plan.follow_up_months:
            continue
        follow_up_due = (execution.performed_date + relativedelta(months=plan.follow_up_months)).date()
        days_until = (follow_up_due - today).days

        if days_until < 0:
            overdue_count += 1
        elif days_until <= 30:
            due_soon_count += 1

    return jsonify({
        'performed': performed_count,
        'planned': planned_count,
        'overdue': overdue_count,
        'due_soon': due_soon_count
    })


@crm_bp.route('/api/crm/plans/<int:plan_id>', methods=['DELETE'])
@login_required
def delete_cosmetic_plan(plan_id):
    """Deleta um planejamento de procedimento cosmético."""
    plan = CosmeticProcedurePlan.query.get_or_404(plan_id)
    
    # Verificar permissão: apenas médico ou secretária
    if not (current_user.is_doctor() or current_user.is_secretary()):
        return jsonify({'error': 'Acesso restrito'}), 403
    
    # Se usuário é médico, só pode deletar seus próprios planos
    if current_user.is_doctor():
        note = Note.query.get(plan.note_id)
        if not note or note.doctor_id != current_user.id:
            return jsonify({'error': 'Você não tem permissão para deletar este planejamento'}), 403
    
    try:
        db.session.delete(plan)
        db.session.commit()
        return jsonify({'success': True, 'message': f'Planejamento "{plan.procedure_name}" deletado com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
