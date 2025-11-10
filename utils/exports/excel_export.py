"""
Exportação de relatórios para Excel (XLSX) usando openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from io import BytesIO

class ExcelExporter:
    """Exportador de relatórios Excel"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_agenda(self, appointments, date_range, doctor_name):
        """
        Exporta agenda para Excel
        
        Args:
            appointments: Lista de agendamentos
            date_range: Tupla (data_inicio, data_fim)
            doctor_name: Nome do médico
            
        Returns:
            BytesIO: Excel gerado
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Agenda"
        
        # Título
        ws['A1'] = f"Agenda Médica - {doctor_name}"
        ws['A1'].font = Font(bold=True, size=14, color="0D6EFD")
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Período: {date_range[0].strftime('%d/%m/%Y')} a {date_range[1].strftime('%d/%m/%Y')}"
        ws.merge_cells('A2:F2')
        
        # Cabeçalhos
        headers = ['Data', 'Hora Início', 'Hora Fim', 'Paciente', 'Status', 'Sala', 'Observações']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Dados
        row = 5
        for apt in appointments:
            ws.cell(row, 1, apt.start_time.strftime('%d/%m/%Y')).border = self.border
            ws.cell(row, 2, apt.start_time.strftime('%H:%M')).border = self.border
            ws.cell(row, 3, apt.end_time.strftime('%H:%M')).border = self.border
            ws.cell(row, 4, apt.patient.name).border = self.border
            ws.cell(row, 5, self._translate_status(apt.status)).border = self.border
            ws.cell(row, 6, apt.room or '-').border = self.border
            ws.cell(row, 7, apt.notes or '-').border = self.border
            row += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 40
        
        # Rodapé
        ws.cell(row + 1, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_surgical_map(self, surgeries, week_start, rooms):
        """
        Exporta mapa cirúrgico semanal para Excel
        
        Args:
            surgeries: Lista de cirurgias
            week_start: Data de início da semana
            rooms: Lista de salas
            
        Returns:
            BytesIO: Excel gerado
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapa Cirúrgico"
        
        # Título
        week_end = week_start + timedelta(days=6)
        ws['A1'] = "Mapa Cirúrgico Semanal"
        ws['A1'].font = Font(bold=True, size=14, color="198754")
        ws.merge_cells('A1:G1')
        
        ws['A2'] = f"Semana de {week_start.strftime('%d/%m/%Y')} a {week_end.strftime('%d/%m/%Y')}"
        ws.merge_cells('A2:G2')
        
        row = 4
        
        for room in rooms:
            room_surgeries = [s for s in surgeries if s.room_id == room.id]
            
            if room_surgeries:
                # Nome da sala
                ws.cell(row, 1, room.name).font = Font(bold=True, size=12)
                ws.merge_cells(f'A{row}:G{row}')
                row += 1
                
                # Cabeçalhos
                headers = ['Data', 'Hora', 'Médico', 'Paciente', 'Procedimento', 'Duração', 'Status']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row, col)
                    cell.value = header
                    cell.fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.border = self.border
                
                row += 1
                
                # Cirurgias
                for surgery in sorted(room_surgeries, key=lambda x: (x.scheduled_date, x.scheduled_time)):
                    ws.cell(row, 1, surgery.scheduled_date.strftime('%d/%m/%Y')).border = self.border
                    ws.cell(row, 2, surgery.scheduled_time.strftime('%H:%M')).border = self.border
                    ws.cell(row, 3, surgery.doctor.name).border = self.border
                    ws.cell(row, 4, surgery.patient_name).border = self.border
                    ws.cell(row, 5, surgery.procedure_type).border = self.border
                    ws.cell(row, 6, f"{surgery.duration_minutes} min").border = self.border
                    ws.cell(row, 7, surgery.status or 'Agendada').border = self.border
                    row += 1
                
                row += 1  # Espaço entre salas
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        
        # Rodapé
        ws.cell(row, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _translate_status(self, status):
        """Traduz status de agendamento"""
        translations = {
            'agendado': 'Agendado',
            'confirmado': 'Confirmado',
            'atendido': 'Atendido',
            'faltou': 'Faltou',
            'cancelado': 'Cancelado'
        }
        return translations.get(status, status.capitalize())
